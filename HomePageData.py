import openpyxl


class HomePageData:

    test_HomePage_data = [{"Firstname": "Naveen", "Email": "mailto:hello@gmail.com", "gender": "Male"},
                            {"Firstname": "Saisha", "Email": "mailto:hello@gmail.com", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\deepak.ms\\AppData\\Local\\Programs\\Python\\Python311\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]